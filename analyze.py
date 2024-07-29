# open -a "Google Chrome" --args --remote-debugging-port=9222
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import time
from pandas_market_calendars import get_calendar
from datetime import datetime, timedelta
import pychrome
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import os
from openpyxl.chart import LineChart, Reference
import json
from utils.limitUtil import getLimitDowns, getLimitUPs

global_wait_seconds = 3

def compare_dates(date1, date2):
    if date1 > date2:
        return 1
    elif date1 < date2:
        return -1
    else:
        return 0

class AnalyzeExit(Exception):
    pass

# 指定开始统计年份
year = 2022
try:
    with open(f'{year}_stocks_data.json', 'r') as file:
        _data = json.load(file)
except FileNotFoundError:
    _data = []
# 创建一个工作簿–
workbook = Workbook()
sheet = workbook.active
# 获取中国交易日历
calendar = get_calendar('XSHG')  # 'XSHG' 表示上海证券交易所的交易日历
# 获取今天的日期
today = datetime.now().date()
# 指定年份的日期范围
# 获取下一个交易日
date_object = datetime.strptime(_data[-1]['date'], '%Y-%m-%d').date()
next_date = calendar.valid_days(start_date=date_object + timedelta(days=1), end_date='2100-01-01')[0]
start_date = next_date.date()
end_date = today
date1 = datetime.strptime(str(start_date), '%Y-%m-%d')
date2 = datetime.strptime(str(end_date), '%Y-%m-%d')
# 比较日期大小
result = compare_dates(date1, date2)
if result == 1:
   raise AnalyzeExit("Terminating analyze.py execution")

# 获取指定日期范围内的工作日日期
workdays = calendar.valid_days(start_date=start_date, end_date=end_date)
# 模拟交易日日期列表
dates = []# 这里替换为你的交易日日期列表
#最高23板，历史记录是20年的斯达半导
limits = ['23','22','21','20','19','18', '17', '16', '15', '14','13','12','11','10','9','8','7','6','5','4','3','2']
for workday in workdays:
    dates.append(str(workday.date()))
# 写入连板数第一列
for idx, limit in enumerate(limits, start=2):
    sheet.cell(row=idx, column=1).value = limit
# 设置第一列所有单元格的字体样式和对齐方式
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.font = Font(size=15, bold=True)  # 设置字体样式
        cell.alignment = Alignment(horizontal='center', vertical='center')  # 设置对齐方式
sheet.column_dimensions['A'].width = 10 
# 调整第一列的高度
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        sheet.row_dimensions[cell.row].height = 30  # 设置第一列的高度为 30

# 设置第一列（除第一个单元格外）的背景色为绿色
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 绿色


# 创建一个Browser实例
browser = pychrome.Browser(url="http://127.0.0.1:9222")
# 新建一个标签页
tab = browser.new_tab()
# 打开链接
tab.start()
tab.Network.enable()
if os.path.exists(f"{year}年A股主板连板数据.xlsx"):
    # 加载工作簿
    workbook = load_workbook(f"{year}年A股主板连板数据.xlsx")
    sheet = workbook['Sheet']
    continue_index = sheet.max_column + 1
else:
    continue_index = 2
# 读取已有的 JSON 文件（如果存在）
try:
    with open(f'{year}_stocks_data.json', 'r') as file:
        existing_data = json.load(file)
except FileNotFoundError:
    existing_data = []

for idx, workday in enumerate(workdays):
    sheet.cell(row=1, column=idx+continue_index).value = str(workday.date())
    # 获取涨停家数
    limit_ups = getLimitUPs(tab,str(workday.date()))
    # 获取跌停家数
    limit_downs = getLimitDowns(tab,str(workday.date()))
    tab.Page.navigate(url="https://www.iwencai.com/unifiedwap/result?w=" + str(workday.date()) + "主板非st涨停,且连板&querytype=stock")
    tab.wait(global_wait_seconds)
    result = tab.Runtime.evaluate(expression="document.documentElement.outerHTML")
    soup = BeautifulSoup(result['result']['value'], 'html.parser')
    # 获取 tbody 中所有的 tr 标签
    tbodies = soup.find_all('tbody', {'data-v-00e1661f': True})
    tbody = tbodies[0]
    if tbody:
       tr_tags = tbody.find_all('tr')
       stocks = []
       # 遍历所有的 tr 标签，并输出其内容
       for tr in tr_tags:
        td_tags = tr.find_all('td')
        if len(td_tags) >= 22:
           #代码
           code_td = td_tags[2] 
           if code_td:
              code = code_td.text.strip() 
           #名字
           name_td = td_tags[3]
           if name_td:
              a_tag = name_td.find('a')
              if a_tag:
                       company_name = a_tag.text
           #价格
           price_td = td_tags[4]
           if price_td:
              price = price_td.text.strip()
           #首次涨停时间
           first_limit_time_td = td_tags[6]
           if first_limit_time_td:
              first_limit_time = first_limit_time_td.text.strip()
           #连板数
           limit_td = td_tags[7]
           if limit_td:
              limit = limit_td.text.strip()
           #最终涨停时间
           final_limit_time_td = td_tags[8]
           if final_limit_time_td:
              final_limit_time = final_limit_time_td.text.strip()
           #涨停概念
           limit_reason_td = td_tags[10]
           if limit_reason_td:
              a_tag = limit_reason_td.find('a')
              if a_tag:
                       limit_reason = a_tag.text
           #涨停封单量
           limit_tocks_td = td_tags[11]
           if limit_tocks_td:
             limit_tocks = limit_tocks_td.text.strip()
           #涨停封单金额
           limit_money_td = td_tags[12]
           if limit_money_td:
             limit_money = limit_money_td.text.strip()
           print(limit_money)
           #涨停封成比
           limit_cheng_ratio_td = td_tags[13]
           if limit_cheng_ratio_td:
             limit_cheng_ratio = limit_cheng_ratio_td.text.strip()
           #涨停封流比
           limit_liu_ratio_td = td_tags[14]
           if limit_liu_ratio_td:
             limit_liu_ratio = limit_liu_ratio_td.text.strip()
           #涨停开板次数
           limit_open_times_td = td_tags[15]
           if limit_open_times_td:
             limit_open_times = limit_open_times_td.text.strip()
           #流通市值
           market_value_td = td_tags[16]
           if market_value_td:
              a_tag = market_value_td.find('a')
              if a_tag:
                       market_value = a_tag.text
           #涨停类型
           limit_type_td = td_tags[18]
           if limit_type_td:
             limit_type = limit_type_td.text.strip()
           #公司注册地址
           company_place_td = td_tags[20]
           if company_place_td:
             company_place = company_place_td.text.strip()
            #公司经营范围
           company_business_td = td_tags[21]
           if company_business_td:
             company_business = company_business_td.text.strip()
        print(limit,limit_ups,limit_downs)
        stock = {'code':code,'name':company_name, 'limit':int(limit),'limit_ups':int(limit_ups),'limit_downs':int(limit_downs),'price':price,'first_limit_time':first_limit_time,
        'final_limit_time':final_limit_time,'limit_reason':limit_reason,'limit_tocks':limit_tocks,'limit_money':limit_money,'limit_cheng_ratio':limit_cheng_ratio,'limit_liu_ratio':limit_liu_ratio,
        'limit_open_times':limit_open_times,'market_value':market_value,'limit_type':limit_type,'company_place':company_place,'company_business':company_business}
        stocks.append(stock)
        sorted_stocks = sorted(stocks, key=lambda x: x['limit'], reverse=True)
        # 将 sorted_stocks 中的股票信息追加到已有数据中
    existing_data.extend([{'date':str(workday.date()),'data':sorted_stocks}])
        # 将数据写入到 JSON 文件中
    with open(f'{year}_stocks_data.json', 'w') as file:
        json.dump(existing_data, file,ensure_ascii=False,  indent=4) 
        # 写入股票名称
    for stock in sorted_stocks:
        print(f"name: {stock['name']}, limit: {stock['limit']}")
        for row in range(23, 1, -1):  # 从13到1倒序遍历第一列的值
            if str(row) == str(stock['limit']):
                target_row = 24-row+1
                start_column = continue_index + idx
                cell = sheet.cell(row=target_row, column=start_column)
                if cell.value:  # 如果单元格有值，则追加新名称和换行符
                   cell.value = f"{cell.value}, {stock['name']}"
                else:
                    sheet.cell(target_row, column=start_column).value = stock['name']
                sheet.cell(target_row, column=start_column).font = Font(size=15, bold=True)  # 设置字号和粗体
                sheet.cell(target_row, column=start_column).alignment = Alignment(horizontal='center', vertical='center')  # 设置居中显示
                break  # 找到匹配值后跳出循环
    for cell in sheet[1]:  # 遍历第一行的所有单元格
        cell.font = Font(size=15, bold=True)  # 设置字体样式
        cell.alignment = Alignment(horizontal='center', vertical='center')  # 设置对齐方式
    # 设置第一行的背景色为红色
    # 设置第一行（除第一个单元格外）的背景色为红色
    for cell in sheet[1][1:]:
        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色
     # 设置列宽自动调整
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in sheet.iter_rows(min_row=1, min_col=col, max_row=sheet.max_row, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                        pass
        adjusted_width = (max_length + 2) * 1.2  # 加一些额外的空间，这个系数可以根据实际情况调整
        sheet.column_dimensions[column_letter].width = adjusted_width
    # 锁定第一列
    sheet.freeze_panes = 'B1'
    # 保存 Excel 文件
    workbook.save(filename=f"{year}年A股主板连板数据.xlsx")