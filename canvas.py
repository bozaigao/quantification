import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
import os
import json

year = 2024
wb = load_workbook(f"{year}年A股主板连板数据.xlsx")
if "连板数据分析" in wb.sheetnames:
    del wb["连板数据分析"]
hart_ws = wb.create_sheet("连板数据分析")
# 获取当前工作表
sheet = wb.active
with open(f'{year}_stocks_data.json', 'r') as file:
        stock_data = json.load(file)
dates = []
for item in stock_data:
    dates.append(item['date'])
if not os.path.exists(f'{year}_dragon_opening_data.json'):
    with open(f'{year}_dragon_opening_data.json', 'w') as file:
        file.write("[]")
with open(f'{year}_dragon_opening_data.json', 'r') as file:
        dragon_opening_data = json.load(file)
# 提取每个日期的最大 limit
max_limits = [max(data['data'], key=lambda x: x['limit'])['limit'] for data in stock_data]
dragon_data = []
for idx, itemData in enumerate(stock_data):
    arr = []
    for item in itemData['data']:
        if item['limit'] == max_limits[idx]:
           arr.append(item)
    filtered_items = {'date':itemData['date'],'data':arr}
    dragon_data.extend([filtered_items])
with open(f'{year}_dragon_data.json', 'w') as file:
    json.dump(dragon_data, file,ensure_ascii=False,  indent=4) 
# 涨停家数
limit_ups = [data['data'][0]['limit_ups'] for data in stock_data]
# 跌停家数
limit_downs = [data['data'][0]['limit_downs'] for data in stock_data]
mapped_data = list(zip(dates, max_limits, limit_ups, limit_downs))
data_dict = {}
# 创建日期和最大 limit 的对应关系字典
for date, max_limit, limit_ups, limit_downs in mapped_data:
    data_dict[date] = {'max_limit': max_limit, 'limit_ups': limit_ups, 'limit_downs': limit_downs}

# 按日期排序
sorted_data = sorted(data_dict.items())
# 写入数据到 Excel 表格
hart_ws.append(['日期', '最大连板高度','涨停家数','跌停家数','最高板数','当日最高溢价',])
for idx, (date, dic) in enumerate(sorted_data):
    limit_downs = dic['limit_downs']
    if len(dragon_opening_data) > 0:
        count = len(dragon_opening_data[idx]['data'])
        percentage_decimal = 0
        for item in dragon_opening_data[idx]['data']:
            if percentage_decimal < float(item['opening_increase'].strip('%')):
               percentage_decimal = float(item['opening_increase'].strip('%'))
        hart_ws.append([date, dic['max_limit'], dic['limit_ups'],limit_downs,count, percentage_decimal])
    else:
        if limit_downs>100:
            limit_downs = 100
        hart_ws.append([date, dic['max_limit'], dic['limit_ups'],limit_downs,0, 0])
# 创建一个 LineChart 对象
chart = LineChart()

# 设置图表的样式和属性
chart.title = "股票最大连板高度趋势图"
chart.x_axis.title = '日期'
chart.y_axis.title = '最大连板高度'

#添加连板趋势分析
data = Reference(hart_ws, min_col=2, min_row=1, max_row=len(dates)+1)
categories = Reference(hart_ws, min_col=1, min_row=2, max_row=len(dates)+1)
# 添加数据到图表
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
# 设置纵坐标范围为最低为2
chart.y_axis.scaling.min = 2
# 设置折线颜色
for series in chart.series:
    series.graphicalProperties.line.solidFill = "0000FF"
# 将图表插入到 Excel 表格中
hart_ws.add_chart(chart, hart_ws.cell(row=1, column=hart_ws.max_column + 1).coordinate)
# 设置图表宽度（根据日期数量自动调整）
chart.width = len(dates)/2

# 涨跌停家数
chart2 = LineChart()
# 设置图表的样式和属性
chart2.title = "股票当日涨跌停数趋势图"
chart2.x_axis.title = '日期'
chart2.y_axis.title = '涨跌停数'
# 创建数据引用对象，包括日期、最大限制、涨停数和跌停数
data2 = Reference(hart_ws, min_col=3, min_row=1, max_col=4, max_row=len(dates)+1)
categories2 = Reference(hart_ws, min_col=1, min_row=2, max_row=len(dates)+1)
# 将数据引用添加到图表中
chart2.add_data(data2, titles_from_data=True)
# 设置图表的类别数据
chart2.set_categories(categories2)
# 设置纵坐标范围为最低为2
chart2.y_axis.scaling.min = 0
# 设置涨停数和跌停数的线条颜色
for series in chart2.series[0:]:  # 从第二个系列开始（第一个系列是日期）
    series.graphicalProperties.line.solidFill = "FF0000"  # 设置为红色
for series in chart2.series[1:]: 
    series.graphicalProperties.line.solidFill = "00FF00"  # 设置为红色
hart_ws.add_chart(chart2, hart_ws.cell(row=16, column=hart_ws.max_column).coordinate)
# # 设置图表宽度（根据日期数量自动调整）
chart2.width = len(dates)/2

if len(dragon_opening_data) > 0:
    # 最高板当日最高溢价
    chart3 = LineChart()
    # 设置图表的样式和属性
    chart3.title = "最高板当日溢价趋势图"
    chart3.x_axis.title = '日期'
    chart3.y_axis.title = '当日溢价'
    # 创建数据引用对象，包括日期、最大限制、涨停数和跌停数
    data3 = Reference(hart_ws, min_col=6, min_row=1, max_row=len(dates)+1)
    categories3 = Reference(hart_ws, min_col=1, min_row=2, max_row=len(dates)+1)
    # 将数据引用添加到图表中
    chart3.add_data(data3, titles_from_data=True)
    # 设置图表的类别数据
    chart3.set_categories(categories3)
    # 设置涨停数和跌停数的线条颜色
    for series in chart3.series[0:]:  # 从第二个系列开始（第一个系列是日期）
        series.graphicalProperties.line.solidFill = "FF0000"  # 设置为红色
    hart_ws.add_chart(chart3, hart_ws.cell(row=31, column=hart_ws.max_column).coordinate)
    # # 设置图表宽度（根据日期数量自动调整）
    chart3.width = len(dates)/2

# 保存 Excel 文件
wb.save(f"{year}年A股主板连板数据.xlsx")
